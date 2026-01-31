"""Export leads to various formats (CSV, JSON, Excel)."""

import csv
import json
import logging
from pathlib import Path
from typing import Optional

from mauritius_lead_scraper.models import Lead

logger = logging.getLogger(__name__)


def export_csv(leads: list[Lead], filepath: str, include_header: bool = True) -> str:
    """Export leads to a CSV file."""
    filepath = _ensure_extension(filepath, ".csv")
    fieldnames = _get_fieldnames()

    with open(filepath, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        if include_header:
            writer.writeheader()
        for lead in leads:
            row = lead.to_dict()
            row["tags"] = "; ".join(row.get("tags", []))
            writer.writerow(row)

    logger.info("Exported %d leads to %s", len(leads), filepath)
    return filepath


def export_json(leads: list[Lead], filepath: str, pretty: bool = True) -> str:
    """Export leads to a JSON file."""
    filepath = _ensure_extension(filepath, ".json")
    data = {
        "total": len(leads),
        "leads": [lead.to_dict() for lead in leads],
    }

    with open(filepath, "w", encoding="utf-8") as f:
        json.dump(data, f, indent=2 if pretty else None, ensure_ascii=False)

    logger.info("Exported %d leads to %s", len(leads), filepath)
    return filepath


def export_excel(leads: list[Lead], filepath: str) -> str:
    """Export leads to an Excel file. Requires openpyxl."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        logger.error("openpyxl is required for Excel export: pip install openpyxl")
        raise ImportError("Install openpyxl for Excel export: pip install openpyxl")

    filepath = _ensure_extension(filepath, ".xlsx")
    wb = Workbook()
    ws = wb.active
    ws.title = "Mauritius Leads"

    # Header styling
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2E86AB", end_color="2E86AB", fill_type="solid")

    fieldnames = _get_fieldnames()
    headers = [_format_header(f) for f in fieldnames]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    # Data rows
    for row_idx, lead in enumerate(leads, 2):
        data = lead.to_dict()
        for col_idx, field in enumerate(fieldnames, 1):
            value = data.get(field, "")
            if isinstance(value, list):
                value = "; ".join(value)
            ws.cell(row=row_idx, column=col_idx, value=value or "")

    # Auto-adjust column widths
    for col in ws.columns:
        max_length = 0
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[col[0].column_letter].width = adjusted_width

    wb.save(filepath)
    logger.info("Exported %d leads to %s", len(leads), filepath)
    return filepath


def _get_fieldnames() -> list[str]:
    """Get the ordered field names for export."""
    return [
        "name",
        "phone",
        "email",
        "website",
        "address",
        "district",
        "category",
        "description",
        "brn",
        "source",
        "tags",
    ]


def _format_header(field: str) -> str:
    """Format a field name as a readable header."""
    special = {"brn": "BRN", "email": "Email", "phone": "Phone"}
    return special.get(field, field.replace("_", " ").title())


def _ensure_extension(filepath: str, ext: str) -> str:
    """Ensure the filepath has the correct extension."""
    p = Path(filepath)
    if p.suffix.lower() != ext.lower():
        return str(p.with_suffix(ext))
    return filepath
